import os
import openpyxl
import xlrd
import xlwt

path = os.environ['HOME'] + '/Documents/a/'
fileList = os.listdir(path)

writeList = []


def readXlsx(excelName):
    workbook = openpyxl.load_workbook(path + excelName)
    sheet = workbook['Sheet1']
    maxCol = sheet.max_column
    write = []
    for t in range(1, 4):
        write.append(sheet.cell(t, 2).value)

    othStr = ''
    for i in range(3, maxCol+1):
        if sheet.cell(1, i).value:
            othStr += sheet.cell(1, i).value + '（'
            for j in range(2, 3):
                othStr += sheet.cell(j, i).value
            othStr += '）\n'

    write.append(othStr)
    writeList.append(write)


def readXls(excelName):
    workbook = xlrd.open_workbook(path + excelName)
    sheets = workbook.sheet_names()
    sheet = workbook.sheet_by_name(sheets[0])
    write = []
    write.append(sheet.cell_value(0, 1))
    writeList.append(write)


def writeXlsx():
    writeFilePath = os.environ['HOME'] + '/Documents/b/write.xlsx'
    print(writeFilePath)
    index = len(writeList)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    #     sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(writeList[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(writeList[i][j]))

    workbook.save(writeFilePath)


for excel in fileList:
    pre = excel.split('.')[1]
    if pre == 'xls':
        readXls(excel)
    else:
        readXlsx(excel)

writeXlsx()
