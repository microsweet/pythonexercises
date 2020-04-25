from PIL import Image
# import pytesseract
import os
import pdfplumber
import re
import openpyxl


# test = pytesseract.image_to_string(Image.open('/home/microsweet/Downloads/登记表1.emf'), lang='chi_sim')

# with open('/home/microsweet/find.txt', 'w') as file_object:
#     file_object.write(test)

pdf_path = os.environ['HOME'] + '/Documents/pdf/'
txt_path = os.environ['HOME'] + '/Documents/txt/'
xls_path = os.environ['HOME'] + '/Documents/xls/'
# path = '/home/microsweet/localNtfs/登记表1.pdf'
pdf_list = os.listdir(pdf_path)
writeRow = []

def pdftotxt():
    for pdf_name in pdf_list:
        pdf = pdfplumber.open(pdf_path+pdf_name)
        for page in pdf.pages:
    #         print(page.extract_text())
            words = page.extract_text()
            txt = txt_path + pdf_name.split('.')[0] + '.txt'
            with open(txt, 'w') as file_object:
                file_object.write(words)
        #     print(words)
        #     arr = words.split('2018 8 13')
        #     print(arr)
        #     for pdf_table in page.extract_tables():
        #         print('---')
        #         table = []
        #         cells = []
        #         for row in pdf_table:
        #             if not any(row):
        #                 # 如果一行全为空，则视为一条记录结束
        #                 if any(cells):
        #                     table.append(cells)
        #                     cells = []
        #             elif all(row):
        #                 # 如果一行全不为空，则本条为新行，上一条结束
        #                 if any(cells):
        #                     table.append(cells)
        #                     cells = []
        #                 table.append(row)
        #             else:
        #                 if len(cells) == 0:
        #                     cells = row
        #                 else:
        #                     for i in range(len(row)):
        #                         if row[i] is not None:
        #                             cells[i] = row[i] if cells[i] is None else cells[i] + row[i]
        #         for row in table:
        #             print('---------------------------')
        #             print([re.sub('\s+', '', cell) if cell is not None else None for cell in row])
        #         print('---------- 分割线 ----------')

        pdf.close()
    txttoexcel()

def txttotxt():
    txt_list = os.listdir(txt_path)
    for txt_name in txt_list:
        with open(txt_path+txt_name) as file_object:
            lines = file_object.readlines()

        wj = []
        for line in lines:
            line = str(line.rstrip())
            wj.append(line)
#         print(str(wj))

        with open(txt_path+txt_name, 'w') as file_object:
                file_object.write(''.join(wj))



    

def txttoexcel():
    txt_list = os.listdir(txt_path)
    for txt_name in txt_list:
        rows = [['姓名', '性别', '身份证号', '籍贯', '毕业中学', '录取学校', '批次']]
        with open(txt_path+txt_name) as file_object:
            article = file_object.read()

#         lines = article.split('2016   7    28')
        lines = article.split('20182018')
        print(txt_name + '：' + str(len(lines)))

        for i in range(0,len(lines)):
            line =lines[i]
            line = line.replace(' ', ' ')
#             print(line)
#             print('-------------------')
            name = re.findall('.*名 (.*) 性.*', line.strip())
            id_card = re.findall('.*身份证号(.*)民　.*', line.strip())
#             gender = re.findall('.*别(.*)应往届.*', line.strip())
#             print(str(name) + str(id_card))
            gender = re.findall('.*别(.*)考生类别.*', line.strip())
#             home = re.findall('.*地(.*)考试类型.*', line.strip())
            home = re.findall('.*地(.*)考试类型.*', line.strip())
            middle_school = re.findall('.*毕业中学(.*)毕业类别.*', line.strip())
            university = re.findall('.*录取到：(.*)学.*', line.strip())
#             print(str(name) + '--' + str(gender) + '---' + str(university) + '---' + str(home))
            
            row = []
            if name:
                try:
                    row.append(name[0].strip())
                    row.append(gender[0].strip())
                    row.append(id_card[0].strip())
                    row.append(home[0].strip())
                    row.append(middle_school[0].strip())
                    if university:
                        row.append(university[0].split('，')[0].strip())
                        row.append(university[0].split('，')[2].strip())
    #                     print(str(university[0].split('，')))
                    else:
                        row.append(str(university).strip())

                    rows.append(row)
                except IndexError:
                    print(str(name))
#     print(rows)
        abc = {'name': txt_name.split('.')[0], 'rows': rows}
        writeRow.append(abc)
#         writeXlsx(txt_name.split('.')[0], rows)

    writeXlsx()


def writeXlsx():
    writeFilePath = xls_path + 'write.xlsx'
    print(writeFilePath)
    workbook = openpyxl.Workbook()
    for j in range(0, len(writeRow)):
#         print(str(obj))
        obj = writeRow[j]
        name = obj['name']
        rows = obj['rows']
        print(name)
        index = len(rows)
#         sheet = workbook.active
        sheet = workbook.create_sheet(title=name, index=j)
        sheet.title = name
        for i in range(0, index):
            for j in range(0, len(rows[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(rows[i][j]))

    workbook.save(writeFilePath)

# pdftotxt()
txttoexcel()
# txttotxt()
