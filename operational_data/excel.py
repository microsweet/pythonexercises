import openpyxl
import os

path = os.environ['HOME'] + '/Documents/xls/2018/'

def read_excel_xlsx():
    num = 0
    txt_list = os.listdir(path)
#     print(str(txt_list))
    for txt_name in txt_list:
        workbook = openpyxl.load_workbook(path+txt_name)
        names = workbook.sheetnames
        for sheet_name in names:
            sheet = workbook[sheet_name]
            maxRow = sheet.max_row - 1
            num += maxRow
    print(num)

read_excel_xlsx()
