#!/usr/bin/env python3
import cx_Oracle
import os
import openpyxl
import xlrd
import shutil


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.load_workbook(path)
    sheet = workbook['Sheet1']
    # sheet.title = sheet_name
    sheet["B2"]=str(value[0])
    sheet["E2"]=str(value[1])
    sheet["B3"]=str(value[2])
    sheet["E3"]=str(value[3])
    sheet["B4"]=str(value[4])
    sheet["E4"]=str(value[5])
    sheet["B5"]=str(value[6])
    sheet["E5"]=str(value[7])
    sheet["B6"]=value[8]*10000
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx():
    # workbook = xlrd.open_workbook(
        # "/home/microsweet/shengchan/单人债权表-7.6交刘凯.xls")
    # sheet = workbook.sheet_by_name("三处")
    # for row in range(3, sheet.nrows-2 ):
        # wpath = "/home/microsweet/shengchan/三处/表三二院欠款项目债券明细表（三处_" + str(sheet.cell_value(row, 0)) + "_" + sheet.cell_value(row, 2) + ").xlsx"
        # arr = [
            # sheet.cell_value(row, 2),
            # sheet.cell_value(row, 0),
            # sheet.cell_value(row, 8),
            # sheet.cell_value(row, 9),
            # sheet.cell_value(row, 1),
            # sheet.cell_value(row, 12),
            # sheet.cell_value(row, 13),
            # sheet.cell_value(row, 14),
            # sheet.cell_value(row, 3)
        # ]
    workbook = openpyxl.load_workbook("/home/microsweet/zq.xlsx")
    sheet = workbook.get_sheet_by_name("Sheet1")
    rows = sheet.max_row
    for row in range(3, rows):
        wpath = "/home/microsweet/shengchan/三处/表三二院欠款项目债券明细表（三处_" + str(sheet.cell_value(row, 0)) + "_" + sheet.cell_value(row, 2) + ").xlsx"
        shutil.copy("/home/microsweet/shengchan/model.xlsx", wpath)
        arr  = [
                sheet.cell(row, 2),
                sheet.cell(row, 0),
                sheet.cell(row, 8),
                sheet.cell(row, 9),
                sheet.cell(row, 1),
                sheet.cell(row, 12),
                sheet.cell(row, 13),
                sheet.cell(row, 14),
                sheet.cell(row, 3)
            ]
        print(sheet.cell_value(row,0))
        # write_excel_xlsx(wpath, 'sheet1', arr)


def query_data():
    read_excel_xlsx()
    # write_excel_xlsx('/home/microsweet/export.xlsx', 'a', rows)
    # print('end')


query_data()
