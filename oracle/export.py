#!/usr/bin/env python3
import cx_Oracle
import os
import openpyxl



def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")
 
 
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()
def query_data():
    os.environ['NSL_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
    #os.environ['NSL_LANG'] = 'AMERICAN_AMERICA.ZHS16GBK'
    print(os.path)

    oracle_tns = 'system/orcl@192.168.128.5:1521/orcl'

    conn = cx_Oracle.connect(oracle_tns)

    curs = conn.cursor()

    sql_path = '/home/microsweet/pythonexercises/oracle/sql.txt'
    with open(sql_path) as file_object:
        contents = file_object.read()

    r = curs.execute(contents)

    rows = []
    for i in r:
        row = list(i)
        rows.append(row)

    for i in rows:
        sql = 'SELECT t.MESSAGE_ FROM GRID_ACTIVIT.ACT_HI_COMMENT t,' \
            ' GRID_ACTIVIT.ACT_HI_TASKINST n ' \
            'WHERE t.TASK_ID_=n.ID_ AND t.PROC_INST_ID_=' \
            + i[1] + \
            ' AND n.NAME_=\'二院处理\' ORDER BY t.TIME_ DESC'
        n = curs.execute(sql)
        for j in n:
            try:
                i.append(j[0])
            except TypeError:
                i.append('')
            break

    curs.close()
    conn.close()
    write_excel_xlsx('/home/microsweet/export.xlsx', 'a', rows)
    print('end')

query_data()
